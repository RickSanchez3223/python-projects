import json
from openpyxl import Workbook, load_workbook
from io import BytesIO, StringIO
from csv import DictReader
from helpers.import_user_validators import ManagerCreateSchema, EmployeeCreateSchema
from marshmallow import ValidationError
from werkzeug.security import generate_password_hash
from helpers.generate_employee_id import generate_employee_id
from datetime import datetime

from constants import MANAGER


def format_json(data):
    return json.dumps(data, indent=4)


def format_csv(data):
    if not data:
        return ""

    # Extract the column names from the first dictionary
    columns = list(data[0].keys())

    # Create a CSV string
    csv_data = ",".join(columns) + "\n"
    for entry in data:
        row = [str(entry[column]) for column in columns]
        csv_data += ",".join(row) + "\n"

    return csv_data


def format_xlsx(data):
    if not data:
        return b""

    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = list(data[0].keys())
    ws.append(headers)

    # Add data to the worksheet
    for entry in data:
        row = [str(entry[column]) for column in headers]
        ws.append(row)

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


def read_csv(file):
    data = []
    file_data = file.read().decode('utf-8')
    csv_data = DictReader(StringIO(file_data))
    for row in csv_data:
        data.append(row)
    return data


def read_xlsx(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    data = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):  # Start from the second row to skip the header row
        data.append({headers[i]: row[i] for i in range(len(headers))})
    return data


def read_json(file):
    return json.load(file)


EXPORT_CONFIG = {
    'json': {
        'content_type': 'application/json',
        'format_function': format_json,
        'read_function': read_json,
    },
    'csv': {
        'content_type': 'text/csv',
        'format_function': format_csv,
        'read_function': read_csv,
    },
    'xlsx': {
        'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'format_function': format_xlsx,
        'read_function': read_xlsx,
    }
}


def get_manager_object(name):
    from app import mongo
    user = mongo.db.user.find_one({'name': name, 'is_deleted': False, 'role': MANAGER})
    return {'id': str(user['_id']), 'name': name}


def validate_data(data):
    results = []
    for row in data:
        print('ROW data: ', row)
        result = {
            "name": row.get('Name', ''),
            "email": row.get('Email', ''),
            "password": generate_password_hash(str(row.get('Password', ''))),
            "role": row.get('Role'),
            "is_admin": False,
            "phone_number": str(row.get('Phone', '')),
            "designation": row.get('Designation', ''),
            "department": row.get('Department', ''),
            "manager": row.get('Manager', None),
            "hired_date": str(row.get('Hired Date', '')),
            "is_deleted": False,
            "employee_id": generate_employee_id(),
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow(),
            "deleted_at": None
        }
        if result['role'] == 'Manager':
            schema = ManagerCreateSchema
            schema_data = {
                "name": result['name'],
                "email": result['email'],
                "phone_number": result['phone_number'],
                "password": result['password'],
                "hired_date": result['hired_date'],
                "role": result['role']
            }
        else:
            schema = EmployeeCreateSchema
            schema_data = {
                "name": result['name'],
                "email": result['email'],
                "phone_number": result['phone_number'],
                "password": result['password'],
                "hired_date": result['hired_date'],
                "designation": result['designation'],
                "department": result['department'],
                "manager": result['manager'],
                "role": result['role']
            }
        try:
            schema().load(schema_data)
        except ValidationError as e:
            return dict(errors=e.messages)
        result['manager'] = get_manager_object(name=result['manager']) if result['manager'] else result['manager']
        results.append(result)

    print('VALIDATION SUCCESS!!', results)
    return dict(data=results)
